import logging
import csv
import io
from typing import List, Dict, Any
from datetime import datetime

logger = logging.getLogger(__name__)

def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Exporter des données en CSV
    
    Args:
        data: Liste de dictionnaires à exporter
        filename: Nom du fichier (optionnel)
    
    Returns:
        str: Contenu CSV
    """
    if not data:
        return ""
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    return output.getvalue()

def export_to_excel(data: List[Dict[str, Any]], filename: str = None):
    """Exporter des données en Excel (nécessite openpyxl)
    
    Args:
        data: Liste de dictionnaires à exporter
        filename: Nom du fichier (optionnel)
    
    Returns:
        bytes: Contenu Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        if not data:
            return None
        
        # Headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Data rows
        for item in data:
            ws.append([item.get(key) for key in headers])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    except ImportError:
        logger.warning("openpyxl not installed, Excel export not available")
        return None
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return None

def sanitize_for_export(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Nettoyer les données avant export
    
    - Convertir les dates en strings
    - Flatten les objets JSONB
    - Masquer les champs sensibles
    
    Args:
        data: Données à nettoyer
    
    Returns:
        List[Dict]: Données nettoyées
    """
    import json
    
    sanitized = []
    
    for item in data:
        clean_item = {}
        for key, value in item.items():
            # Convertir dates
            if isinstance(value, datetime):
                clean_item[key] = value.isoformat()
            # Flatten JSONB
            elif isinstance(value, (dict, list)):
                clean_item[key] = json.dumps(value)
            # Masquer champs sensibles
            elif key in ['password', 'mfa_secret', 'secret_key']:
                clean_item[key] = '***'
            else:
                clean_item[key] = value
        
        sanitized.append(clean_item)
    
    return sanitized
